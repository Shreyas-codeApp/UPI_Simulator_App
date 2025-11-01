import streamlit as st
import openpyxl
import pandas as pd


wb = openpyxl.load_workbook('database.xlsx')
sheet = wb.active


if "account" not in st.session_state:
    st.session_state["account"] = {}
if "validation" not in st.session_state:
    st.session_state["validation"] = 0
if "show_interface" not in st.session_state:
    st.session_state["show_interface"] = False

st.set_page_config(page_title="UPI Simulator", page_icon="🏦", layout="wide")


def create_account(username, PIN):
    
    new_row = sheet.max_row + 1
    sheet.cell(row=new_row, column=1).value = username
    sheet.cell(row=new_row, column=2).value = PIN
    sheet.cell(row=new_row, column=3).value = 0.0
    wb.save('database.xlsx')
    st.success(f"Account created as {username}.")
    st.info("Please log in to your account from the login section.", icon="ℹ️")

def login_account(username, PIN):
    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == username and str(sheet.cell(row=i, column=2).value) == str(PIN):
            return True
    return False

def set_balance(amount):
    st.session_state["account"]["balance"] = float(amount)
    sheet.cell(row=st.session_state["account"]["index"], column=3).value = float(amount)
    wb.save('database.xlsx')
    st.toast(f"Balance set to {amount}", icon="✅", duration=3)
    st.success(f"Set balance to {amount}")

def make_transaction(amount, recipient):
    amount = float(amount)
    current_balance = float(st.session_state["account"]["balance"])
    if amount > current_balance:
        st.error("Transaction failed: Insufficient balance.", icon="🚫")
        st.info("Please press 'Pay!' again to exit transaction mode.", icon="ℹ️")
        
        return False
    else:
        st.session_state["account"]["balance"] = current_balance - amount
        sheet.cell(row=st.session_state["account"]["index"], column=3).value = st.session_state["account"]["balance"]
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == recipient:
                recipient_balance = sheet.cell(row=i, column=3).value
                recipient_balance = float(recipient_balance) if recipient_balance else 0.0
                sheet.cell(row=i, column=3).value = recipient_balance + amount
                break
        wb.save("database.xlsx")
        st.balloons()
        st.toast("Transaction successful!", icon="✅", duration=3)
        st.success("Payment successful!")
        st.info("Please press 'Pay!' again to exit transaction mode.", icon="ℹ️")
        return True


def expense_entry(attribute_col, data):
    row = st.session_state["account"]["index"]
    existing = sheet.cell(row=row, column=attribute_col).value
    if existing:
        new_val = str(existing) + "," + str(data)
    else:
        new_val = str(data)
    sheet.cell(row=row, column=attribute_col).value = new_val
    wb.save("database.xlsx")

def deposit_entry(attribute_col, data, recipient_username=None):
    
    target = recipient_username or st.session_state["Transactions"]["recipient"]
    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == target:
            existing = sheet.cell(row=i, column=attribute_col).value
            if existing:
                sheet.cell(row=i, column=attribute_col).value = str(existing) + "," + str(data)
            else:
                sheet.cell(row=i, column=attribute_col).value = str(data)
            wb.save("database.xlsx")
            break
def remove_contact(contact_name):
    contacts = st.session_state["account"].get("all_contacts", [])
    if not contact_name:
        st.toast("No contact selected.", icon="⚠️", duration=3)
        return
    if contact_name in contacts:
        contacts.remove(contact_name)
        st.session_state["account"]["all_contacts"] = contacts
        sheet.cell(row=st.session_state["account"]["index"], column=4).value = ",".join(contacts) if contacts else ""
        wb.save('database.xlsx')
        st.toast(f"Contact '{contact_name}' removed successfully.", icon="✅", duration=4)
    else:
        st.toast(f"Contact '{contact_name}' not found in your list.", icon="⚠️", duration=4)




with st.sidebar:
    with st.expander("Accounts", expanded=False):
        
        with st.expander("Sign up", expanded=False):
            st.write("Create a new account")
            st.caption("Choose a unique username and a secure 4-digit PIN.")
            new_username = st.text_input("New Username").strip()
            new_PIN = st.text_input("Set PIN", type="password").strip()
            signup = st.button("Sign Up")
            if signup:
                if not new_username:
                    st.error("Username cannot be empty.")
                elif new_username.isspace():
                    st.error("Username cannot be spaces only.")
                elif new_username.isnumeric():
                    st.error("Username cannot be purely numeric.")
                elif any(not c.isalnum() for c in new_username):
                    st.error("Username cannot contain special characters.")
                else:
                    banned_keywords = {
                        "usrname","username","user","admin","administrator","root","system","password",
                        "contacts","upi","test","guest","null","none","default","bank","balance","money",
                        "transaction","developer","account","login","signup"
                    }
                    if any(word in new_username.lower() for word in banned_keywords):
                        st.error("This username is not allowed. Please choose a different username.")
                    elif not new_PIN:
                        st.error("Enter a 4-digit PIN.")
                    elif not new_PIN.isdigit() or len(new_PIN) != 4:
                        st.error("PIN must be a 4-digit number.")
                    else:
                       
                        existing = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row+1)]
                        if new_username in existing:
                            st.error("Username already exists. Please choose a different username.")
                        else:
                            create_account(new_username, new_PIN)

        
        with st.expander("Login", expanded=False):
            st.write("Log in to your account")
            st.caption("Enter your username and 4-digit PIN to log in to your account.")
            input_username = st.text_input("Username").strip()
            input_PIN = st.text_input("PIN", type="password").strip()
            login_btn = st.button("Login")
            if login_btn and (input_username and input_PIN) and st.session_state["validation"] < 3:
                if login_account(input_username, input_PIN):
                    
                    st.session_state["account"] = {}
                    st.success(f"Logged in as {input_username}.")
                    st.session_state["account"]["username"] = input_username
                    st.session_state["account"][input_username] = input_PIN  
                    st.session_state["validation"] = 0

                    
                    for i in range(2, sheet.max_row + 1):
                        if sheet.cell(row=i, column=1).value == input_username:
                            st.session_state["account"]["index"] = i
                            break

                   
                    balance_val = sheet.cell(row=st.session_state["account"]["index"], column=3).value
                    st.session_state["account"]["balance"] = float(balance_val) if balance_val else 0.0

                   
                    contacts_cell = sheet.cell(row=st.session_state["account"]["index"], column=4).value
                    if contacts_cell:
                        
                        st.session_state["account"]["all_contacts"] = [c.strip() for c in str(contacts_cell).split(",") if c.strip()]
                    else:
                        st.session_state["account"]["all_contacts"] = []

                    st.session_state["show_interface"] = True
                    successful_login = True
                else:
                    
                    found_user = any(sheet.cell(row=i, column=1).value == input_username for i in range(2, sheet.max_row + 1))
                    if not found_user:
                        st.error("Username not found in the database. Please sign up first.")
                    else:
                        st.session_state["validation"] += 1
                        remaining = 3 - st.session_state["validation"]
                        if remaining > 0:
                            st.error("Incorrect PIN. Please try again.", icon="🚫")
                            st.warning(f"Warning: {remaining} login attempts remaining.", icon="⚠️")
                        else:
                            st.error("Session terminated due to multiple failed login attempts.", icon="🚫")
                            st.info("Please refresh the page to try again.", icon="ℹ️")

        
        if "account" in st.session_state and st.session_state["account"].get("username"):
            if st.session_state["show_interface"]:
                with st.expander("Account Settings", expanded=False):
                    with st.expander("Change PIN", expanded=False):
                        st.write("Want to change your PIN?")
                        new_PIN_change = st.text_input("Enter new PIN", type="password").strip()
                        change_PIN = st.button("Change PIN")
                        if change_PIN:
                            if not new_PIN_change:
                                st.error("Enter a new 4-digit PIN.")
                            elif not new_PIN_change.isdigit() or len(new_PIN_change) != 4:
                                st.error("PIN must be a 4-digit number.")
                            else:
                                sheet.cell(row=st.session_state["account"]["index"], column=2).value = new_PIN_change
                                wb.save('database.xlsx')
                                st.success("PIN changed successfully.")

                    with st.expander("Delete Account", expanded=False):
                        st.warning("Warning: This action is irreversible.", icon="⚠️")
                        delete_account = st.button("Delete Account")
                        if delete_account:
                            username_to_delete = st.session_state["account"]["username"]
                            
                            for i in range(2, sheet.max_row + 1):
                                if sheet.cell(row=i, column=1).value == username_to_delete:
                                    sheet.delete_rows(i)
                                    wb.save('database.xlsx')
                                    st.session_state.clear()
                                    st.success("Account deleted successfully. Please refresh.")
                                    st.stop()

    with st.expander("About", expanded=False):
        st.markdown("""
            ### 📘 About  
            Explore how digital payments work — safely and interactively.  
            The **UPI Simulator App** lets you experience UPI-style transactions in a fully offline environment.  

            **Key Features:**  
            - 🧾 **Create & manage virtual accounts** – Sign up, log in, and personalize your mock UPI wallet.  
            - 💸 **Simulate secure transactions** – Send or receive payments using virtual balances.  
            - 📊 **Track expenses & categories** – Review spending patterns and payment history with ease.  
            - 👥 **Add contacts & manage transfers** – Keep your own digital contact list for smooth mock payments.  
        """)
        st.warning("This app doesn't involve transaction with real money", icon="⚠️")


st.title(":rainbow[UPI Simulator App]")
st.caption("Simulate • Learn • Master Digital Payments")

if st.session_state.get("show_interface") is False:
    st.info("Please log in to your account to continue. If you don't have an account, please sign up first.\nBe sure to read the About section.", icon="ℹ️")
    st.info("Please click on the '>>' button to find the account and the about section")

if st.session_state.get("show_interface"):
    account_details, balance, contacts, transactions, expenses = st.tabs(
        ["Account details", "Balance", "Contacts", "Transactions", "Expenses"]
    )

    
    with contacts:
        st.subheader("Add and manage your contacts here.")
        show_contacts,add_contact,remove_contact_tab = st.columns(3)

        
        with add_contact:
            new_contact = st.text_input("Enter contact name to add").strip()
            add_contact = st.button("Add Contact")
            current_username = st.session_state["account"]["username"]
            if add_contact:
                if not new_contact:
                    st.error("Enter a contact name.")
                elif new_contact == current_username:
                    st.error("You cannot add yourself as a contact.")
                else:
                    
                    contact_exists_in_db = any(sheet.cell(row=i, column=1).value == new_contact for i in range(2, sheet.max_row + 1))
                    if not contact_exists_in_db:
                        st.error("Could not find the user in database.")
                    else:
                       
                        if new_contact in st.session_state["account"]["all_contacts"]:
                            st.error("Contact already exists.")
                        else:
                            
                            st.session_state["account"]["all_contacts"].append(new_contact)
                            
                            sheet.cell(row=st.session_state["account"]["index"], column=4).value = ",".join(st.session_state["account"]["all_contacts"])
                            wb.save('database.xlsx')
                            st.success(f"Contact {new_contact} added.")
        with remove_contact_tab:
            remove_contact_name = st.selectbox("Select contact to remove", st.session_state["account"]["all_contacts"])
            remove_contact_btn = st.button("Remove Contact")
            if remove_contact_btn:
                if not remove_contact_name:
                    st.error("Select a contact to remove.")
                else:
                    remove_contact(remove_contact_name)
        with show_contacts:

            st.write("Your Contacts:")
            contacts_cell = sheet.cell(row=st.session_state["account"]["index"], column=4).value
            if contacts_cell:
                
                valid_users = [sheet.cell(row=j, column=1).value for j in range(2, sheet.max_row + 1)]
                current_contacts = [c.strip() for c in st.session_state["account"]["all_contacts"] if c.strip()]
                removed = [c for c in current_contacts if c not in valid_users]
                if removed:
                    st.toast(f"Removed {len(removed)} invalid contact(s): {', '.join(removed)}", icon="⚠️", duration=5)
                    current_contacts = [c for c in current_contacts if c not in removed]
                    st.session_state["account"]["all_contacts"] = current_contacts
                    sheet.cell(row=st.session_state["account"]["index"], column=4).value = ",".join(current_contacts)
                    wb.save('database.xlsx')

                
                if st.session_state["account"]["all_contacts"]:
                    st.write(", ".join(st.session_state["account"]["all_contacts"]))
                else:
                    st.info("No contacts added yet.")
            else:
                st.info("No contacts added yet. Please add contacts in the contacts tab.", icon="ℹ️")

    
    with transactions:
        st.subheader("Make a Transaction")
        if "Transactions" not in st.session_state:
            st.session_state["Transactions"] = {
                "recipient": "",
                "amount": 0.0,
                "category": "",
                "note": "",
                "payment_state": "show_recipient_menu"
            }

        if st.session_state["account"]["all_contacts"]:
            select_recipient, select_amount, details, payment_summary = st.tabs(
                ["Select Recipient", "Enter Amount", "Details", "Payment Summary"]
            )

            with select_recipient:
                recipient = st.selectbox("Select Recipient", st.session_state["account"]["all_contacts"])
                if st.button("Select Recipient") and recipient:
                    st.session_state["Transactions"]["recipient"] = recipient
                    st.success("Recipient selected. Move to Enter Amount.")
                    st.session_state["Transactions"]["payment_state"] = "enter_amount"

            with select_amount:
                if st.session_state["Transactions"]["payment_state"] == "enter_amount":
                    amount = st.number_input("Enter amount", min_value=0.0, value=0.0)
                    amount_button = st.button("Select amount")
                    if amount_button:
                        if amount > 0:
                            st.session_state["Transactions"]["amount"] = float(amount)
                            st.success("Amount selected. Move to Details.")
                            st.session_state["Transactions"]["payment_state"] = "get_details"
                        else:
                            st.error("Enter amount more than zero")
                else:
                    st.info("Please complete the previous step(s)!")

            with details:
                categories = [
                    "Food & Dining", "Groceries", "Shopping", "Bills & Utilities",
                    "Travel", "Entertainment", "Health & Fitness", "Education",
                    "Transfers", "Contributions"
                ]
                if st.session_state["Transactions"]["payment_state"] == "get_details":
                    get_note = st.text_input("Enter a note (optional):")
                    get_note = get_note.strip()
                    if st.button("Add note (optional)"):
                        if "," in get_note:
                            st.error("Note shouldn't have comma(,)", icon = "🚫")
                        elif get_note.isdigit():
                            st.error("Note shouldn't be purely numeric", icon = "🚫")
                        else:
                            st.session_state["Transactions"]["note"] = get_note
                            st.success("Note saved.", icon = "✅" )
                    select_category = st.selectbox("Select category:", categories)
                    if st.button("Select category"):
                        st.session_state["Transactions"]["category"] = select_category
                        st.success("Category selected.")
                        st.session_state["Transactions"]["payment_state"] = "confirm_transaction"
                else:
                    st.info("Please complete the previous step(s)!")

            with payment_summary:
                txn = st.session_state["Transactions"]
                if txn["payment_state"] == "confirm_transaction" and txn["recipient"] and txn["amount"] > 0:
                    st.markdown("### 💸 Transaction Summary")
                    st.markdown(f"**Recipient:** :blue[{txn['recipient']}]  \n**Amount:** :green[${txn['amount']}]  \n**Category:** :violet[{txn['category']}]  \n**Note:** :orange[{txn['note'] or 'No note added'}]")
                   
                    if st.button("Pay!"):
                        before_balance = st.session_state["account"]["balance"]
                        make_transaction(txn["amount"], txn["recipient"])
                        after_balance = st.session_state["account"]["balance"]

                        if after_balance < before_balance:  # Only proceed if transaction succeeded
                            expense_entry(5, txn["amount"])
                            expense_entry(6, txn["category"])
                            expense_entry(7, txn["note"])
                            expense_entry(8, txn["recipient"])
                            expense_entry(9, "-")

                            deposit_entry(5, txn["amount"], recipient_username=txn["recipient"])
                            deposit_entry(6, txn["category"], recipient_username=txn["recipient"])
                            deposit_entry(7, txn["note"], recipient_username=txn["recipient"])
                            deposit_entry(8, "-", recipient_username=txn["recipient"])
                            deposit_entry(9, st.session_state["account"]["username"], recipient_username=txn["recipient"])

                            st.session_state["Transactions"] = {
                                "recipient": "",
                                "amount": 0.0,
                                "category": "",
                                "note": "",
                                "payment_state": "show_recipient_menu"
                            }
                        else:
                            st.session_state["Transactions"]["payment_state"] = "show_recipient_menu"
                            
                else:
                    st.info("Please complete the previous step(s)!")

        else:
            st.info("No contacts available. Please add contacts in the contacts tab to make a transaction.", icon="ℹ️")

    
    with expenses:
        if "expenses" not in st.session_state:
            st.session_state["expenses"] = {}

        row_index = st.session_state["account"]["index"]

        def split_cell(col):
            v = sheet.cell(row=row_index, column=col).value
            if not v:
                return []
            return [s.strip() for s in str(v).split(",")]

        amounts = split_cell(5)
        categories = split_cell(6)
        notes = split_cell(7)
        recipients = split_cell(8)
        senders = split_cell(9)

        st.session_state["expenses"]["amounts"] = amounts
        st.session_state["expenses"]["categories"] = categories
        st.session_state["expenses"]["notes"] = notes
        st.session_state["expenses"]["recipients"] = recipients
        st.session_state["expenses"]["senders"] = senders

        if not (amounts or categories or notes or recipients or senders):
            st.write("No transactions yet!")
        else:
            n = max(len(amounts), len(categories), len(notes), len(recipients), len(senders))
            def pad(lst): return lst + [""] * (n - len(lst))
            amounts, categories, notes, recipients, senders = map(pad, [amounts, categories, notes, recipients, senders])

            def to_number(s):
                try:
                    return float(s)
                except:
                    return s
            amounts = [to_number(x) for x in amounts]

            rows = []
            for i in range(n):
                rows.append({
                    "Amount": f"${amounts[i]:,.2f}",
                    "Category": categories[i],
                    "Note": notes[i],
                    "Recipient": recipients[i],
                    "Sender": senders[i]
                })

            df = pd.DataFrame(rows)

            
            st.subheader("Track Expenses")
            query = st.text_input("🔍 Search expenses (any field)")
            if query:
                df = df[df.apply(lambda row: row.astype(str).str.contains(query, case=False).any(), axis=1)]

            
            sort_col = st.selectbox("Sort by", df.columns, index=0)
            sort_order = st.radio("Order", ["Ascending", "Descending"], horizontal=True)

            df_sorted = df.sort_values(
                by=sort_col,
                ascending=(sort_order == "Ascending"),
                ignore_index=True
            )

            st.dataframe(df_sorted, use_container_width=True)

    
    with balance:
        st.subheader("Balance Section")
        default_balance = float(sheet.cell(row=st.session_state["account"]["index"], column=3).value or 0.0)
        balance_amount = st.slider("Set your balance", 0.0, 10000.0, default_balance, 10.0)
        if st.button("Set Balance"):
            set_balance(balance_amount)
        
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == st.session_state["account"]["username"]:
                st.session_state["account"]["balance"] = float(sheet.cell(row=i, column=3).value or 0.0)
                break
        st.write("Current Balance: $", st.session_state["account"]["balance"])

    
    with account_details:
        keys = st.session_state["account"].keys()
        display_username = st.session_state["account"]["username"]
        st.subheader(f"Welcome, {display_username}")
        balance_col, contacts_col, info_col = st.columns(3)
        with balance_col:
            st.write("Your Balance:")
            st.write("$", st.session_state["account"]["balance"])
            if st.session_state["account"]["balance"] == 0.0:
                st.info("Please set a balance in the balance tab", icon="⚠️")
        with contacts_col:
            st.write("Your Contacts:")
            cell_val = sheet.cell(row=st.session_state["account"]["index"], column=4).value
            if cell_val:
                st.write(cell_val)
            else:
                st.info("No contacts added yet. Please add contacts in the contacts tab.", icon="ℹ️")

    logout_button = st.button("Logout")
    if logout_button:
        st.session_state["show_interface"] = False
        st.session_state["account"] = {}
        st.session_state["Transactions"] = {"payment_state": "enter_amount"}
        st.session_state["expenses"] = {}
        st.rerun()



